from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from mock_dbs.setup_databases import setup_all
from mock_dbs.setup_mongo import setup_analytics_mongo
from mock_dbs.setup_bar import setup_bar
from schema_registry import (
    get_full_registry, build_schema_context,
    is_read_query, is_schema_modifying_query, execute_query,
    register_dynamic_service, register_mongo_service, BASE_DIR,
)
from claude_service import generate_sql, generate_create_db

# SQLite files always start with this 16-byte header
SQLITE_MAGIC = b"SQLite format 3\x00"

app = FastAPI(title="QueryBuddy API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize on startup
registry = {}
schema_context = ""

@app.on_event("startup")
def startup():
    global registry, schema_context
    # SQLite mock services
    setup_all()
    # Bar/restaurant mock service (MongoDB — requires local mongod)
    bar_client = setup_bar()
    if bar_client:
        register_mongo_service(
            service_name="bar_service",
            client=bar_client,
            db_name="bar",
            description="Bar & restaurant — drinks, food, employees, tabs, inventory, shifts (MongoDB)",
        )
    # MongoDB mock service (in-memory via mongomock — no server needed)
    mongo_client = setup_analytics_mongo()
    if mongo_client:
        register_mongo_service(
            service_name="analytics_service",
            client=mongo_client,
            db_name="analytics",
            description="User behaviour analytics — events and sessions (MongoDB)",
        )
    registry = get_full_registry()
    schema_context = build_schema_context(registry)
    print("QueryBuddy ready.")

# ---- Models ----

class Message(BaseModel):
    role: str  # "user" or "assistant"
    content: str

class QueryRequest(BaseModel):
    message: str
    history: Optional[List[Message]] = []

class ExecuteRequest(BaseModel):
    service: str
    sql: str

class MongoConnectRequest(BaseModel):
    connection_string: str   # e.g. "mongodb://localhost:27017"
    db_name: str             # database to expose
    service_name: Optional[str] = None  # auto-derived if omitted

class CreateDbRequest(BaseModel):
    description: str  # plain-English description of the desired database

# ---- Routes ----

@app.get("/api/health")
def health():
    return {"status": "ok"}

@app.get("/api/schema")
def get_schema():
    return registry

MAX_MESSAGE_LENGTH = 4000  # characters

@app.post("/api/query")
def query(req: QueryRequest):
    # Bug 6 fix: validate message length and content
    msg = req.message.strip()
    if not msg:
        raise HTTPException(status_code=400, detail="Message cannot be empty.")
    if len(msg) > MAX_MESSAGE_LENGTH:
        raise HTTPException(
            status_code=400,
            detail=f"Message too long ({len(msg)} chars). Maximum is {MAX_MESSAGE_LENGTH}."
        )

    history = [{"role": m.role, "content": m.content} for m in req.history]

    # Schema context is now always injected via the system prompt in
    # generate_sql (Bug 2 fix), so we no longer need the if/else split here.
    result = generate_sql(msg, schema_context, history)

    # Auto-execute read queries and attach results.
    # Short-circuit: if multiple queries are chained (stitching_note is set)
    # and the first query returns 0 rows, skip the rest — they would filter
    # on IDs that don't exist and return misleading data.
    queries = result.get("queries", [])
    is_chained = bool(result.get("stitching_note"))
    upstream_empty = False

    for i, q in enumerate(queries):
        q["is_read"] = is_read_query(q.get("sql", ""))

        if is_chained and upstream_empty and i > 0:
            q["skipped"] = True
            q["skip_reason"] = (
                "Skipped: the upstream query returned 0 rows, so this query "
                "would produce no meaningful results."
            )
            continue

        if q["is_read"]:
            q["result"] = execute_query(q["service"], q["sql"])
            # Mark upstream empty only for the first query in a chain
            if i == 0 and is_chained:
                rows = q["result"].get("rows")
                if rows is not None and len(rows) == 0:
                    upstream_empty = True

    return result


@app.post("/api/create-db")
def create_db(req: CreateDbRequest):
    """
    Create a new SQLite database from a plain-English description.
    Claude generates the schema (CREATE TABLE) and seed data (INSERT).
    """
    global registry, schema_context
    import sqlite3

    desc = req.description.strip()
    if not desc:
        raise HTTPException(status_code=400, detail="Description cannot be empty.")
    if len(desc) > MAX_MESSAGE_LENGTH:
        raise HTTPException(status_code=400, detail="Description too long.")

    # Ask Claude to generate the SQL
    result = generate_create_db(desc)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])

    db_name = re.sub(r'[^a-z0-9]+', '_', result.get("db_name", "custom_db").lower()).strip('_')
    description = result.get("description", f"User-created database: {db_name}")
    statements = result.get("sql_statements", [])

    if not statements:
        raise HTTPException(status_code=400, detail="Claude did not generate any SQL statements.")

    # Derive unique service name
    base_service = db_name + "_service"
    service_name = base_service
    counter = 2
    while service_name in registry:
        service_name = f"{db_name}_{counter}_service"
        counter += 1

    # Create the SQLite database and execute all statements
    db_path = os.path.join(BASE_DIR, f"{service_name}.db")
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        for stmt in statements:
            cursor.execute(stmt)
        conn.commit()
        conn.close()
    except Exception as e:
        # Clean up the file on failure
        if os.path.exists(db_path):
            os.remove(db_path)
        raise HTTPException(status_code=400, detail=f"SQL execution failed: {e}")

    # Register and rebuild schema
    register_dynamic_service(service_name, db_path)
    registry = get_full_registry()
    schema_context = build_schema_context(registry)

    return {
        "service_name": service_name,
        "description": description,
        "schema": registry[service_name],
        "statements_executed": len(statements),
    }


@app.post("/api/connect-mongo")
def connect_mongo(req: MongoConnectRequest):
    """
    Connect to a real (or mongomock) MongoDB instance, register it as a live
    service, and rebuild the schema context so Claude knows about its collections.
    """
    global registry, schema_context

    try:
        import pymongo
        client = pymongo.MongoClient(req.connection_string, serverSelectionTimeoutMS=5000)
        # Force a connection attempt to surface errors early
        client.server_info()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot connect to MongoDB: {e}")

    # Derive a unique service name
    base_name = re.sub(r'[^a-z0-9]+', '_', req.db_name.lower()).strip('_')
    base_service = (req.service_name or base_name) + "_service"
    service_name = base_service
    counter = 2
    while service_name in registry:
        service_name = f"{base_name}_{counter}_service"
        counter += 1

    register_mongo_service(
        service_name=service_name,
        client=client,
        db_name=req.db_name,
        description=f"MongoDB: {req.db_name} @ {req.connection_string}",
    )
    registry = get_full_registry()
    schema_context = build_schema_context(registry)

    return {"service_name": service_name, "schema": registry[service_name]}


@app.post("/api/upload-db")
async def upload_db(file: UploadFile = File(...)):
    """
    Accept a drag-and-dropped database file, save/convert it to SQLite,
    register it as a live service, and rebuild the schema context.
    Supports: .db / .sqlite / .sqlite3 (native SQLite) and .xlsx / .xls (Excel).
    """
    global registry, schema_context

    filename = file.filename or ""
    is_sqlite = bool(re.search(r'\.(db|sqlite|sqlite3)$', filename, re.IGNORECASE))
    is_excel  = bool(re.search(r'\.xlsx?$', filename, re.IGNORECASE))

    if not is_sqlite and not is_excel:
        raise HTTPException(
            status_code=400,
            detail="Only .db, .sqlite, .sqlite3, .xlsx, or .xls files are accepted."
        )

    content = await file.read()

    # ── Derive a unique service name from the filename ────────────────────────
    stem = re.sub(r'\.(db|sqlite|sqlite3|xlsx|xls)$', '', filename, flags=re.IGNORECASE)
    base_name = re.sub(r'[^a-z0-9]+', '_', stem.lower()).strip('_')
    base_service = base_name + "_service"

    service_name = base_service
    counter = 2
    while service_name in registry:
        service_name = f"{base_name}_{counter}_service"
        counter += 1

    db_path = os.path.join(BASE_DIR, f"{service_name}.db")

    if is_sqlite:
        # ── Native SQLite: validate magic bytes and save directly ─────────────
        if not content.startswith(SQLITE_MAGIC):
            raise HTTPException(
                status_code=400,
                detail="File does not appear to be a valid SQLite database."
            )
        with open(db_path, "wb") as f:
            f.write(content)

    elif is_excel:
        # ── Excel: convert each sheet into a SQLite table ─────────────────────
        import sqlite3, io, tempfile
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="openpyxl is not installed. Run: pip install openpyxl"
            )

        try:
            wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {e}")

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        tables_created = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                continue  # skip empty sheets or sheets with only headers

            # First row = column headers
            raw_headers = rows[0]
            headers = []
            for i, h in enumerate(raw_headers):
                col_name = re.sub(r'[^a-z0-9_]+', '_', str(h or f"col_{i}").lower()).strip('_')
                if not col_name:
                    col_name = f"col_{i}"
                headers.append(col_name)

            # Sanitize table name
            table_name = re.sub(r'[^a-z0-9_]+', '_', sheet_name.lower()).strip('_') or "sheet"

            # Infer column types from first few data rows
            col_types = []
            for ci in range(len(headers)):
                sample_vals = [rows[ri][ci] for ri in range(1, min(len(rows), 20)) if ci < len(rows[ri])]
                sample_vals = [v for v in sample_vals if v is not None]
                if all(isinstance(v, (int,)) for v in sample_vals) and sample_vals:
                    col_types.append("INTEGER")
                elif all(isinstance(v, (int, float)) for v in sample_vals) and sample_vals:
                    col_types.append("REAL")
                else:
                    col_types.append("TEXT")

            col_defs = ", ".join(f'"{h}" {col_types[i]}' for i, h in enumerate(headers))
            cursor.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({col_defs})')

            placeholders = ", ".join("?" * len(headers))
            for row in rows[1:]:
                # Pad or trim row to match header count
                padded = list(row[:len(headers)])
                while len(padded) < len(headers):
                    padded.append(None)
                # Convert values to strings for TEXT columns, keep native for others
                cleaned = []
                for ci, val in enumerate(padded):
                    if val is None:
                        cleaned.append(None)
                    elif col_types[ci] == "TEXT":
                        cleaned.append(str(val))
                    else:
                        cleaned.append(val)
                cursor.execute(f'INSERT INTO "{table_name}" VALUES ({placeholders})', cleaned)
            tables_created += 1

        conn.commit()
        conn.close()
        wb.close()

        if tables_created == 0:
            if os.path.exists(db_path):
                os.remove(db_path)
            raise HTTPException(
                status_code=400,
                detail="No sheets with data found in the Excel file."
            )

    # ── Register & rebuild schema ─────────────────────────────────────────────
    register_dynamic_service(service_name, db_path)
    registry = get_full_registry()
    schema_context = build_schema_context(registry)

    return {
        "service_name": service_name,
        "schema": registry[service_name],
    }


@app.post("/api/execute")
def execute(req: ExecuteRequest):
    # Prompt 17 fix: allow user-approved DML (INSERT/UPDATE/DELETE) but reject
    # schema-modifying DDL (DROP/CREATE/ALTER/TRUNCATE) which is irreversible.
    # The original Bug 7 fix was too broad — it blocked all non-SELECT including
    # legitimate UPDATE statements the user explicitly clicked EXECUTE on.
    if is_schema_modifying_query(req.sql):
        raise HTTPException(
            status_code=400,
            detail=(
                "Schema-modifying operations (DROP, CREATE, ALTER, TRUNCATE) "
                "are not permitted via this endpoint."
            )
        )
    return execute_query(req.service, req.sql)
